from __future__ import annotations

from dataclasses import dataclass, field
from typing import Dict, List, Tuple, Optional
import numpy as np
from openpyxl import Workbook

from .inputs import InputParameters

EPS = 1e-30


class FormulaError(ValueError):
    pass


def parse_formula(formula: str) -> Dict[str, int]:
    s = formula.replace(" ", "")
    i = 0

    def parse_group() -> Dict[str, int]:
        nonlocal i
        counts: Dict[str, int] = {}
        while i < len(s):
            if s[i] == "(":
                i += 1
                inner = parse_group()
                if i >= len(s) or s[i] != ")":
                    raise FormulaError(f"Unmatched '(' in {formula}")
                i += 1
                mult = parse_int()
                for el, n in inner.items():
                    counts[el] = counts.get(el, 0) + n * mult
            elif s[i] == ")":
                break
            else:
                el = parse_element()
                n = parse_int()
                counts[el] = counts.get(el, 0) + n
        return counts

    def parse_element() -> str:
        nonlocal i
        if i >= len(s) or not s[i].isalpha() or not s[i].isupper():
            raise FormulaError(f"Expected element at position {i} in {formula}")
        el = s[i]
        i += 1
        while i < len(s) and s[i].isalpha() and s[i].islower():
            el += s[i]
            i += 1
        return el

    def parse_int() -> int:
        nonlocal i
        if i >= len(s) or not s[i].isdigit():
            return 1
        j = i
        while j < len(s) and s[j].isdigit():
            j += 1
        val = int(s[i:j])
        i = j
        return val

    counts = parse_group()
    if i != len(s):
        raise FormulaError(f"Could not parse full formula {formula} (stopped at {i})")
    return counts


@dataclass
class ComponentRegistry:
    atomic_weights: Dict[str, float] = field(default_factory=dict)
    mw: Dict[str, float] = field(default_factory=dict)
    _order: List[str] = field(default_factory=list)
    frozen: bool = False

    def add_species(self, name: str, mw_g_mol: Optional[float] = None, formula: Optional[str] = None) -> None:
        if name in self.mw:
            return
        if self.frozen:
            raise ValueError(f"Registry frozen; cannot add species '{name}'")
        if mw_g_mol is None:
            if formula is None:
                raise ValueError(f"Need mw_g_mol or formula for species '{name}'")
            mw_g_mol = self.mw_from_formula(formula)
        self.mw[name] = float(mw_g_mol)
        self._order.append(name)

    def mw_from_formula(self, formula: str) -> float:
        atoms = parse_formula(formula)
        mw = 0.0
        for el, n in atoms.items():
            if el not in self.atomic_weights:
                raise ValueError(f"Atomic weight for element '{el}' not provided (needed for {formula})")
            mw += self.atomic_weights[el] * n
        return mw

    @property
    def species(self) -> Tuple[str, ...]:
        return tuple(self._order)

    def index(self) -> Dict[str, int]:
        return {sp: i for i, sp in enumerate(self._order)}

    def n(self) -> int:
        return len(self._order)

    def freeze(self) -> None:
        self.frozen = True


@dataclass
class Stream:
    name: str
    reg: ComponentRegistry
    mol: Dict[str, float] = field(default_factory=dict)
    T: float = 298.15
    p: float = 100000.0
    phase: str = "L"
    producer: Optional[str] = None

    def __post_init__(self):
        if self.T <= 0:
            raise ValueError(f"{self.name}: T must be > 0 K")
        if self.p <= 0:
            raise ValueError(f"{self.name}: p must be > 0 Pa")
        if not self.phase:
            raise ValueError(f"{self.name}: phase must be non-empty")

        cleaned: Dict[str, float] = {}
        for sp, v in self.mol.items():
            if v < 0:
                raise ValueError(f"{self.name}: negative flow for {sp}: {v}")
            if abs(v) > EPS:
                cleaned[sp] = float(v)
                if sp not in self.reg.mw:
                    self.reg.add_species(sp, mw_g_mol=1.0)
        self.mol = cleaned

    def get(self, sp: str) -> float:
        return self.mol.get(sp, 0.0)

    def set(self, sp: str, v: float) -> None:
        if v < 0:
            raise ValueError(f"{self.name}: negative flow for {sp}: {v}")
        if sp not in self.reg.mw:
            self.reg.add_species(sp, mw_g_mol=1.0)
        if abs(v) <= EPS:
            self.mol.pop(sp, None)
        else:
            self.mol[sp] = float(v)

    def total_molar_flow(self) -> float:
        return float(sum(self.mol.values()))

    def to_dense(self) -> np.ndarray:
        x = np.zeros(self.reg.n(), dtype=float)
        idx = self.reg.index()
        for sp, v in self.mol.items():
            x[idx[sp]] = v
        return x

    def from_dense(self, x: np.ndarray) -> None:
        idx = self.reg.index()
        for sp in self.reg.species:
            v = float(x[idx[sp]])
            if v < 0.0 and abs(v) <= 1e-12:
                v = 0.0
            self.set(sp, v)

    def mass_flow_g_s(self) -> float:
        m = 0.0
        for sp, n in self.mol.items():
            m += n * self.reg.mw.get(sp, np.nan)
        return float(m)


class UnitOp:
    def __init__(self, name: str):
        self.name = name
        self.inlets: Dict[str, Stream] = {}
        self.outlets: Dict[str, Stream] = {}

    def add_inlet(self, port: str, s: Stream) -> None:
        self.inlets[port] = s

    def add_outlet(self, port: str, s: Stream) -> None:
        self.outlets[port] = s
        s.producer = self.name

    def apply(self) -> None:
        raise NotImplementedError


class Flowsheet:
    def __init__(self, reg: ComponentRegistry, *, default_T: float, default_p: float):
        self.reg = reg
        self.default_T = float(default_T)
        self.default_p = float(default_p)
        self.streams: Dict[str, Stream] = {}
        self.units: Dict[str, UnitOp] = {}

    def add_stream(self, s: Stream) -> Stream:
        if s.name in self.streams:
            raise ValueError(f"Stream '{s.name}' already exists")
        self.streams[s.name] = s
        return s

    def new_process_stream(self, name: str, *, phase: str, mol: Optional[Dict[str, float]] = None) -> Stream:
        return self.add_stream(Stream(
            name=name, reg=self.reg, mol=(mol or {}),
            T=self.default_T, p=self.default_p, phase=phase
        ))

    def new_exogenous_stream(self, name: str, *, mol: Optional[Dict[str, float]] = None) -> Stream:
        spec = InputParameters.EXOGENOUS_STREAM_SPECS.get(name)
        if spec is None:
            raise ValueError(f"Missing EXOGENOUS_STREAM_SPECS for '{name}'")
        return self.add_stream(Stream(
            name=name, reg=self.reg, mol=(mol or {}),
            T=float(spec["T"]), p=float(spec["p"]), phase=str(spec["phase"])
        ))

    def add_unit(self, unit: UnitOp) -> UnitOp:
        if unit.name in self.units:
            raise ValueError(f"Unit '{unit.name}' already exists")
        self.units[unit.name] = unit
        return unit


def export_to_excel(fs: Flowsheet, path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Streams"
    ws.append(["stream", "T [K]", "p [Pa]", "phase", "F_total [mol/s]", "m_dot [g/s]"])
    for sname in sorted(fs.streams.keys()):
        s = fs.streams[sname]
        ws.append([s.name, s.T, s.p, s.phase, s.total_molar_flow(), s.mass_flow_g_s()])

    ws2 = wb.create_sheet("MolarFlows")
    sp = list(fs.reg.species)
    ws2.append(["stream"] + sp)
    for sname in sorted(fs.streams.keys()):
        s = fs.streams[sname]
        ws2.append([s.name] + [s.get(x) for x in sp])

    ws3 = wb.create_sheet("MoleFractions")
    ws3.append(["stream"] + sp)
    for sname in sorted(fs.streams.keys()):
        s = fs.streams[sname]
        Ft = s.total_molar_flow()
        ws3.append([s.name] + [(s.get(x) / Ft) if Ft > EPS else 0.0 for x in sp])

    wb.save(path)


def stamp_outlets_to_unit_conditions(unit: UnitOp, unit_conditions: Dict[str, Dict[str, float]]) -> None:
    cond = unit_conditions.get(unit.name)
    if not cond:
        return
    T = cond.get("T", None)
    p = cond.get("p", None)
    for s in unit.outlets.values():
        if T is not None:
            s.T = float(T)
        if p is not None:
            s.p = float(p)

def _needs_T(src, T: float | None) -> bool:
    return (T is not None) and (abs(float(src.T) - float(T)) > 1e-9)


def _needs_p(src, p: float | None) -> bool:
    return (p is not None) and (abs(float(src.p) - float(p)) > 1e-6)


def snapshot_streams(fs: Flowsheet) -> dict[str, tuple[np.ndarray, float, float]]:
    snap: dict[str, tuple[np.ndarray, float, float]] = {}
    for name, s in fs.streams.items():
        snap[name] = (s.to_dense().copy(), float(s.T), float(s.p))
    return snap


def max_rel_change_all_streams(
    prev: dict[str, tuple[np.ndarray, float, float]],
    curr: dict[str, tuple[np.ndarray, float, float]],
    *,
    eps: float = 1e-12,
    include_Tp: bool = True,
) -> float:
    worst = 0.0
    for name, (x0, T0, p0) in prev.items():
        x1, T1, p1 = curr[name]
        denom = max(float(np.max(np.abs(x0))), eps)
        dx = float(np.max(np.abs(x1 - x0))) / denom
        worst = max(worst, dx)
        if include_Tp:
            dT = abs(T1 - T0) / max(abs(T0), eps)
            dp = abs(p1 - p0) / max(abs(p0), eps)
            worst = max(worst, float(dT), float(dp))
    return float(worst)


def set_aqueous_molarities(
    s: Stream,
    *,
    Vdot_L_s: float,
    molarity: Dict[str, float],
    water_name: str = "H2O",
    water_molarity_pure: float = 55.5,
) -> None:
    if Vdot_L_s < 0:
        raise ValueError("Vdot_L_s must be >= 0")
    n_sol = 0.0
    for sp, M in molarity.items():
        n = M * Vdot_L_s
        s.set(sp, n)
        n_sol += n
    s.set(water_name, max(water_molarity_pure * Vdot_L_s - n_sol, 0.0))


def compute_required_acid_purge_fraction(
    s23: Stream,
    *,
    target_hno3_mol_s: float,
    target_h2o_mol_s: float,
    f_min: float = 0.0,
    f_max: float = 0.999,
) -> float:
    def req_f(n_s23: float, n_target: float) -> float:
        if n_s23 <= EPS:
            return 0.0
        if n_target >= n_s23:
            return 0.0
        return 1.0 - (n_target / n_s23)

    f_hno3 = req_f(s23.get("HNO3"), target_hno3_mol_s)
    f_h2o = req_f(s23.get("H2O"), target_h2o_mol_s)
    f = max(f_hno3, f_h2o)
    return float(min(max(f, f_min), f_max))


def size_makeup_for_targets(
    recycle: Stream,
    makeup: Stream,
    *,
    target_hno3_mol_s: float,
    target_h2o_mol_s: float,
) -> None:
    if target_hno3_mol_s < 0 or target_h2o_mol_s < 0:
        raise ValueError("targets must be >= 0")
    rec_hno3 = recycle.get("HNO3")
    rec_h2o = recycle.get("H2O")
    make_hno3 = max(target_hno3_mol_s - rec_hno3, 0.0)
    make_h2o = max(target_h2o_mol_s - rec_h2o, 0.0)
    makeup.mol = {}
    makeup.set("HNO3", make_hno3)
    makeup.set("H2O", make_h2o)


def size_solvent_makeup_for_required_org_flow(
    reg: ComponentRegistry,
    recycle: Stream,
    makeup: Stream,
    *,
    n_org_req_total: float,
    tbp_wt: float,
    tbp_name: str = "TBP",
    dil_name: str = "Dodecane",
) -> None:
    M_tbp = reg.mw[tbp_name]
    M_dil = reg.mw[dil_name]
    M_mix = 1.0 / (tbp_wt / M_tbp + (1.0 - tbp_wt) / M_dil)
    mdot_req = n_org_req_total * M_mix
    mdot_tbp_req = tbp_wt * mdot_req
    mdot_dil_req = (1.0 - tbp_wt) * mdot_req

    mdot_tbp_rec = recycle.get(tbp_name) * M_tbp
    mdot_dil_rec = recycle.get(dil_name) * M_dil
    mdot_tbp_make = max(mdot_tbp_req - mdot_tbp_rec, 0.0)
    mdot_dil_make = max(mdot_dil_req - mdot_dil_rec, 0.0)

    makeup.mol = {}
    makeup.set(tbp_name, mdot_tbp_make / max(M_tbp, EPS))
    makeup.set(dil_name, mdot_dil_make / max(M_dil, EPS))


def _safe_sheet_name(name: str) -> str:
    # Excel: <= 31 chars, no : \ / ? * [ ]
    bad = [":", "\\", "/", "?", "*", "[", "]"]
    for b in bad:
        name = name.replace(b, "_")
    name = name.strip()
    if not name:
        name = "Sheet"
    return name[:31]


def _is_scalar(x: Any) -> bool:
    return isinstance(x, (int, float, str, bool)) or x is None


def _summarize_value(v: Any) -> str:
    # Make something Excel-friendly (single cell)
    if _is_scalar(v):
        return "" if v is None else str(v)
    if isinstance(v, (list, tuple, set)):
        return ", ".join(str(x) for x in list(v))
    if isinstance(v, dict):
        # short dict representation
        items = list(v.items())
        return "; ".join(f"{k}:{val}" for k, val in items)
    return str(v)


def export_units_to_excel(
    fs: Flowsheet,
    params: Any,
    path: str,
    *,
    unit_conditions: Dict[str, Dict[str, float]] | None = None,
) -> None:
    """
    Create an Excel workbook with 1 sheet per unit/vessel.
    Each sheet includes:
      - unit metadata (name, class)
      - operating conditions (T/p from unit_conditions)
      - input parameters (all InputParameters attributes matching unit prefix like 'R1_' or 'X1_')
      - calculated parameters (unit attributes, incl. e.g. N_used/OA_design/OA_used if present)
      - inlet/outlet stream summaries
    """
    unit_conditions = unit_conditions or getattr(params, "UNIT_CONDITIONS", {}) or {}

    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)

    # Gather InputParameters attrs once
    # (Skip dunders/callables)
    param_items: Dict[str, Any] = {}
    for k in dir(params):
        if k.startswith("__"):
            continue
        v = getattr(params, k)
        if callable(v):
            continue
        param_items[k] = v

    for uname, u in fs.units.items():
        ws = wb.create_sheet(_safe_sheet_name(uname))

        # -----------------------
        # Header / metadata
        # -----------------------
        ws.append(["Unit", uname])
        ws.append(["Class", u.__class__.__name__])
        cond = unit_conditions.get(uname, {})
        ws.append(["Condition_T [K]", cond.get("T", "")])
        ws.append(["Condition_p [Pa]", cond.get("p", "")])
        ws.append([])

        # -----------------------
        # InputParameters (prefix match)
        # -----------------------
        # prefix = "R1_" from "R1_Dissolver" etc.
        prefix = uname.split("_", 1)[0] + "_"
        ws.append(["InputParameters (matched prefix)", prefix])
        ws.append(["param", "value"])
        # deterministic ordering
        for k in sorted(param_items.keys()):
            if k.startswith(prefix):
                ws.append([k, _summarize_value(param_items[k])])
        ws.append([])

        # -----------------------
        # Calculated / unit attributes
        # -----------------------
        ws.append(["Calculated / Unit attributes", ""])
        ws.append(["attr", "value"])
        # include only "simple" / useful attributes; skip inlets/outlets objects themselves
        skip = {"inlets", "outlets", "name"}
        for k in sorted(getattr(u, "__dict__", {}).keys()):
            if k in skip:
                continue
            v = getattr(u, k)
            # Avoid dumping Stream objects or registry-heavy stuff
            if hasattr(v, "to_dense") or hasattr(v, "reg"):
                continue
            ws.append([k, _summarize_value(v)])
        ws.append([])

        # -----------------------
        # Inlets / outlets summaries
        # -----------------------
        ws.append(["Inlets", ""])
        ws.append(["port", "stream", "phase", "T [K]", "p [Pa]", "F_total [mol/s]", "m_dot [g/s]"])
        for port, s in u.inlets.items():
            ws.append([port, s.name, s.phase, s.T, s.p, s.total_molar_flow(), s.mass_flow_g_s()])
        ws.append([])

        ws.append(["Outlets", ""])
        ws.append(["port", "stream", "phase", "T [K]", "p [Pa]", "F_total [mol/s]", "m_dot [g/s]"])
        for port, s in u.outlets.items():
            ws.append([port, s.name, s.phase, s.T, s.p, s.total_molar_flow(), s.mass_flow_g_s()])

    wb.save(path)

# ==============================================================================
# RECYCLE / TEAR-STREAM SOLVER (ROBUST)
# ==============================================================================

from dataclasses import dataclass
from typing import Callable, Iterable


@dataclass
class RecycleResult:
    converged: bool
    iterations: int
    error: float
    error_history: list[float]


def _pack_tears(
    fs: Flowsheet,
    tears: list[str],
    *,
    include_Tp: bool,
) -> np.ndarray:
    """Concatenate tear stream dense molar vectors (+ optional T/p) into one vector."""
    chunks: list[np.ndarray] = []
    for nm in tears:
        s = fs.streams[nm]
        chunks.append(s.to_dense())
        if include_Tp:
            chunks.append(np.array([float(s.T), float(s.p)], dtype=float))
    return np.concatenate(chunks) if chunks else np.zeros((0,), dtype=float)


def _unpack_tears(
    fs: Flowsheet,
    tears: list[str],
    x: np.ndarray,
    *,
    include_Tp: bool,
) -> None:
    """Write concatenated tear vector back into fs.streams tear objects."""
    i = 0
    nsp = fs.reg.n()
    for nm in tears:
        s = fs.streams[nm]
        s.from_dense(x[i:i+nsp])
        i += nsp
        if include_Tp:
            s.T = float(x[i]); s.p = float(x[i+1])
            i += 2


def _rel_error(x_new: np.ndarray, x_old: np.ndarray, *, eps: float = 1e-12) -> float:
    denom = max(float(np.max(np.abs(x_old))), eps)
    return float(np.max(np.abs(x_new - x_old))) / denom


class _Anderson:
    """
    Small Anderson acceleration for fixed-point x = F(x).
    Stores last m residuals and does a regularized least squares.
    """
    def __init__(self, m: int = 6, lam: float = 1e-10):
        self.m = int(m)
        self.lam = float(lam)
        self.X: list[np.ndarray] = []
        self.R: list[np.ndarray] = []

    def propose(self, x_fp: np.ndarray, x_prev: np.ndarray) -> np.ndarray:
        """
        Given fixed-point output x_fp = F(x_prev), form a mixed iterate.
        Residual r = x_fp - x_prev.
        """
        r = (x_fp - x_prev)
        self.X.append(x_prev.copy())
        self.R.append(r.copy())
        if len(self.X) > self.m:
            self.X.pop(0)
            self.R.pop(0)

        # Not enough history -> just return fixed-point
        k = len(self.R)
        if k < 2:
            return x_fp

        # Solve min ||R c|| s.t. sum(c)=1, then x = sum(c_i * (X_i + R_i)) = sum(c_i * F(X_i))
        # Build matrix of residuals
        Rm = np.column_stack(self.R)  # (n, k)
        RtR = Rm.T @ Rm + self.lam * np.eye(k)

        ones = np.ones((k, 1))
        KKT = np.block([[RtR, ones],
                        [ones.T, np.zeros((1, 1))]])
        rhs = np.concatenate([np.zeros((k, 1)), np.ones((1, 1))], axis=0)

        try:
            sol = np.linalg.solve(KKT, rhs)
            c = sol[:k, 0]
        except np.linalg.LinAlgError:
            return x_fp

        # Weighted combination of past fixed-point images: F(X_i) = X_i + R_i
        Fm = np.column_stack([self.X[j] + self.R[j] for j in range(k)])
        x_and = Fm @ c
        return x_and


def solve_recycles(
    fs: Flowsheet,
    *,
    tears: Iterable[str],
    evaluate_once: Callable[[], None],
    max_iter: int = 200,
    tol: float = 1e-8,
    relax: float = 0.5,
    include_Tp: bool = True,
    use_anderson: bool = True,
    anderson_m: int = 6,
    verbose: bool = False,
) -> RecycleResult:
    """
    Robust recycle loop solving x = F(x) where x is the tear-stream state.

    - evaluate_once(): runs the flowsheet "once" (your unit_sequence + any sizing).
      It must update the tear stream(s) as part of the run.

    Convergence is based on tear-stream change only.
    """
    tears = list(tears)
    if not tears:
        # Nothing to solve; just run once
        evaluate_once()
        return RecycleResult(True, 1, 0.0, [0.0])

    for nm in tears:
        if nm not in fs.streams:
            raise KeyError(f"Tear stream '{nm}' not found in flowsheet.streams")

    anderson = _Anderson(m=anderson_m) if use_anderson else None

    # Initial x
    x_prev = _pack_tears(fs, tears, include_Tp=include_Tp)
    hist: list[float] = []

    for it in range(1, max_iter + 1):
        # Run one pass -> updates fs.streams[...] including tear(s)
        evaluate_once()

        x_fp = _pack_tears(fs, tears, include_Tp=include_Tp)  # fixed-point image F(x_prev)
        err = _rel_error(x_fp, x_prev)
        hist.append(err)

        if verbose:
            print(f"[recycle] it={it:4d} err={err:.3e}")

        if err < tol:
            return RecycleResult(True, it, err, hist)

        # Propose next x using Anderson (optional)
        if anderson is not None:
            x_prop = anderson.propose(x_fp, x_prev)
        else:
            x_prop = x_fp

        # Under-relaxation (always applied)
        x_next = (1.0 - relax) * x_prev + relax * x_prop

        # Write back the tear(s) for next iteration
        _unpack_tears(fs, tears, x_next, include_Tp=include_Tp)

        # update
        x_prev = x_next

    # Not converged
    return RecycleResult(False, max_iter, hist[-1] if hist else float("inf"), hist)

# -----------------------------------------------------------------------------
# Deterministic / idempotent conditioning support
# -----------------------------------------------------------------------------

def _cond_key(src_name: str, unit_name: str, T: float | None, p: float | None) -> tuple:
    # Keys must be hashable and stable
    return (str(src_name), str(unit_name),
            None if T is None else float(T),
            None if p is None else float(p))


def get_or_create_conditioned_stream(
    fs: Flowsheet,
    *,
    src: Stream,
    unit_name: str,
    target_T: float | None,
    target_p: float | None,
    make_HX: callable,
    make_PC: callable,
) -> Stream:
    """
    Returns a deterministic conditioned stream for feeding `unit_name`.

    - Creates at most one HX and one PC (and at most one output stream per step)
      for the (src.name, unit_name, target_T, target_p) signature.
    - Subsequent calls reuse the previously created objects.
    """
    # Attach a cache lazily
    if not hasattr(fs, "_conditioning_cache"):
        fs._conditioning_cache = {}  # type: ignore[attr-defined]

    cache: dict = fs._conditioning_cache  # type: ignore[attr-defined]
    key = _cond_key(src.name, unit_name, target_T, target_p)

    if key in cache:
        return cache[key]

    out = src

    # HX step (if needed)
    if target_T is not None and abs(float(out.T) - float(target_T)) > 1e-9:
        hx_name = f"HX__{src.name}__to__{unit_name}"
        sT_name = f"{src.name}__to__{unit_name}__T"

        # stream
        if sT_name in fs.streams:
            sT = fs.streams[sT_name]
        else:
            sT = fs.new_process_stream(sT_name, phase=out.phase, mol={})
        sT.phase = out.phase
        sT.T = float(target_T)
        sT.p = float(out.p)

        # unit
        if hx_name in fs.units:
            hx = fs.units[hx_name]
        else:
            hx = make_HX(hx_name, float(target_T))
            fs.add_unit(hx)

        hx.inlets["in"] = out
        hx.outlets["out"] = sT
        sT.producer = hx.name

        out = sT

    # PC step (if needed)
    if target_p is not None and abs(float(out.p) - float(target_p)) > 1e-6:
        pc_name = f"PC__{src.name}__to__{unit_name}"
        sp_name = f"{src.name}__to__{unit_name}__p"

        # stream
        if sp_name in fs.streams:
            sp = fs.streams[sp_name]
        else:
            sp = fs.new_process_stream(sp_name, phase=out.phase, mol={})
        sp.phase = out.phase
        sp.p = float(target_p)
        sp.T = float(out.T)

        # unit
        if pc_name in fs.units:
            pc = fs.units[pc_name]
        else:
            pc = make_PC(pc_name, float(target_p))
            fs.add_unit(pc)

        pc.inlets["in"] = out
        pc.outlets["out"] = sp
        sp.producer = pc.name

        out = sp

    cache[key] = out
    return out
