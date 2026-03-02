"""
=============================================================================
SIMPLIFIED PROCESS FLOWSHEET — CORE FRAMEWORK
=============================================================================
Provides:
  - ComponentRegistry  : species MW database
  - Stream             : molar flow + T/p/phase
  - UnitOp             : base class for process units
  - Flowsheet          : container + helpers
  - Anderson / solve_recycles : robust recycle convergence
  - export_to_excel    : stream table + unit summary
=============================================================================
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple

import numpy as np

EPS = 1e-30


# ─────────────────────────────────────────────────────────────────────────────
# FORMULA PARSER + MOLECULAR WEIGHT
# ─────────────────────────────────────────────────────────────────────────────

def _parse_formula(formula: str) -> Dict[str, int]:
    """Return element: count dict from a chemical formula string."""
    s = formula.replace(" ", "")
    i = 0

    def parse_group() -> Dict[str, int]:
        nonlocal i
        counts: Dict[str, int] = {}
        while i < len(s):
            if s[i] == "(":
                i += 1
                inner = parse_group()
                if i >= len(s) or s[i] != ")":
                    raise ValueError(f"Unmatched '(' in {formula!r}")
                i += 1
                mult = _parse_int()
                for el, n in inner.items():
                    counts[el] = counts.get(el, 0) + n * mult
            elif s[i] == ")":
                break
            else:
                el = _parse_element()
                n = _parse_int()
                counts[el] = counts.get(el, 0) + n
        return counts

    def _parse_element() -> str:
        nonlocal i
        if i >= len(s) or not s[i].isupper():
            raise ValueError(f"Expected element at pos {i} in {formula!r}")
        el = s[i]; i += 1
        while i < len(s) and s[i].islower():
            el += s[i]; i += 1
        return el

    def _parse_int() -> int:
        nonlocal i
        if i >= len(s) or not s[i].isdigit():
            return 1
        j = i
        while j < len(s) and s[j].isdigit():
            j += 1
        val = int(s[i:j]); i = j
        return val

    result = parse_group()
    if i != len(s):
        raise ValueError(f"Could not parse full formula {formula!r}")
    return result


def mw_from_formula(formula: str, atomic_weights: Dict[str, float]) -> float:
    atoms = _parse_formula(formula)
    mw = 0.0
    for el, n in atoms.items():
        if el not in atomic_weights:
            raise ValueError(f"Atomic weight not known for element '{el}' (in {formula!r})")
        mw += atomic_weights[el] * n
    return mw


# ─────────────────────────────────────────────────────────────────────────────
# COMPONENT REGISTRY
# ─────────────────────────────────────────────────────────────────────────────

class ComponentRegistry:
    """
    Database of species and their molecular weights.
    Populated once at flowsheet construction; thereafter essentially read-only.
    """

    def __init__(self, atomic_weights: Optional[Dict[str, float]] = None):
        self._aw: Dict[str, float] = dict(atomic_weights or {})
        self._mw: Dict[str, float] = {}
        self._order: List[str] = []

    # ── Registration ─────────────────────────────────────────────────────────

    def register(
        self,
        name: str,
        *,
        mw: Optional[float] = None,
        formula: Optional[str] = None,
    ) -> None:
        """Register a species.  Idempotent — safe to call multiple times."""
        if name in self._mw:
            return
        if mw is not None:
            self._mw[name] = float(mw)
        elif formula is not None:
            self._mw[name] = mw_from_formula(formula, self._aw)
        else:
            raise ValueError(f"Must supply mw or formula for species '{name}'")
        self._order.append(name)

    def ensure(self, name: str) -> None:
        """Register a species with MW=1 if not already known (fallback)."""
        if name not in self._mw:
            self._mw[name] = 1.0
            self._order.append(name)

    # ── Access ────────────────────────────────────────────────────────────────

    @property
    def species(self) -> Tuple[str, ...]:
        return tuple(self._order)

    def n(self) -> int:
        return len(self._order)

    def index(self) -> Dict[str, int]:
        return {sp: i for i, sp in enumerate(self._order)}

    def get_mw(self, name: str) -> float:
        return self._mw.get(name, 1.0)


# ─────────────────────────────────────────────────────────────────────────────
# STREAM
# ─────────────────────────────────────────────────────────────────────────────

class Stream:
    """
    A process stream carrying molar flows of multiple species
    plus temperature, pressure, and phase label.
    """

    def __init__(
        self,
        name: str,
        reg: ComponentRegistry,
        *,
        mol: Optional[Dict[str, float]] = None,
        T: float = 298.15,
        p: float = 1e5,
        phase: str = "L",
    ):
        self.name = name
        self.reg = reg
        self.T = float(T)
        self.p = float(p)
        self.phase = phase
        self.mol: Dict[str, float] = {}
        self.producer: Optional[str] = None   # name of UnitOp that fills this stream

        for sp, v in (mol or {}).items():
            self.set(sp, v)

    # ── Accessors ─────────────────────────────────────────────────────────────

    def get(self, sp: str) -> float:
        return self.mol.get(sp, 0.0)

    def set(self, sp: str, v: float) -> None:
        v = float(v)
        if v < -1e-12:
            raise ValueError(f"{self.name}: negative flow {v} for '{sp}'")
        if v < 0.0:
            v = 0.0
        self.reg.ensure(sp)
        if v < EPS:
            self.mol.pop(sp, None)
        else:
            self.mol[sp] = v

    def add(self, sp: str, v: float) -> None:
        self.set(sp, self.get(sp) + v)

    def clear(self) -> None:
        self.mol.clear()

    # ── Aggregate ─────────────────────────────────────────────────────────────

    def total_molar_flow(self) -> float:
        return float(sum(self.mol.values()))

    def mass_flow_g_s(self) -> float:
        return float(sum(n * self.reg.get_mw(sp) for sp, n in self.mol.items()))

    # ── Dense array I/O (for recycle solver) ─────────────────────────────────

    def to_dense(self) -> np.ndarray:
        idx = self.reg.index()
        x = np.zeros(self.reg.n(), dtype=float)
        for sp, v in self.mol.items():
            if sp in idx:
                x[idx[sp]] = v
        return x

    def from_dense(self, x: np.ndarray) -> None:
        idx = self.reg.index()
        for sp in self.reg.species:
            v = float(x[idx[sp]])
            self.set(sp, max(v, 0.0))

    def copy_from(self, other: "Stream") -> None:
        """Copy mol dict (and T/p/phase) from another stream."""
        self.mol = {sp: v for sp, v in other.mol.items()}
        self.T = other.T
        self.p = other.p
        self.phase = other.phase

    def __repr__(self) -> str:
        F = self.total_molar_flow()
        return f"Stream({self.name!r}, phase={self.phase}, T={self.T:.1f} K, F={F:.4f} mol/s)"


# ─────────────────────────────────────────────────────────────────────────────
# UNIT OP  (base class)
# ─────────────────────────────────────────────────────────────────────────────

class UnitOp:
    def __init__(self, name: str):
        self.name = name
        self.inlets: Dict[str, Stream] = {}
        self.outlets: Dict[str, Stream] = {}
        self.calcs: Dict[str, Any] = {}   # computed quantities stored here

    def connect_inlet(self, port: str, s: Stream) -> None:
        self.inlets[port] = s

    def connect_outlet(self, port: str, s: Stream) -> None:
        self.outlets[port] = s
        s.producer = self.name

    def apply(self) -> None:
        raise NotImplementedError(f"{self.__class__.__name__}.apply() not implemented")

    def _stamp_conditions(self, T: Optional[float], p: Optional[float]) -> None:
        """Stamp T/p onto all outlet streams."""
        for s in self.outlets.values():
            if T is not None:
                s.T = T
            if p is not None:
                s.p = p


# ─────────────────────────────────────────────────────────────────────────────
# CONDITIONER  (HeaterCooler / PressureChanger)
# ─────────────────────────────────────────────────────────────────────────────

class Conditioner(UnitOp):
    """
    Adjusts T and/or p of a stream without changing composition.
    Created automatically by the flowsheet when CONDITION_FEEDS is True.
    """

    def __init__(self, name: str, *, target_T: Optional[float] = None, target_p: Optional[float] = None):
        super().__init__(name)
        self.target_T = target_T
        self.target_p = target_p

    def apply(self) -> None:
        src = self.inlets["in"]
        dst = self.outlets["out"]
        # Copy composition
        dst.mol = {sp: v for sp, v in src.mol.items()}
        dst.phase = src.phase
        dst.T = self.target_T if self.target_T is not None else src.T
        dst.p = self.target_p if self.target_p is not None else src.p


# ─────────────────────────────────────────────────────────────────────────────
# FLOWSHEET
# ─────────────────────────────────────────────────────────────────────────────

class Flowsheet:
    """
    Container for streams and units.  Tracks unit execution order.
    """

    def __init__(self, reg: ComponentRegistry, *, default_T: float = 298.15, default_p: float = 1e5):
        self.reg = reg
        self.default_T = default_T
        self.default_p = default_p
        self.streams: Dict[str, Stream] = {}
        self.units: Dict[str, UnitOp] = {}
        self._unit_sequence: List[UnitOp] = []

    def add_stream(self, name: str, *, phase: str = "L",
                   mol: Optional[Dict[str, float]] = None,
                   T: Optional[float] = None, p: Optional[float] = None) -> Stream:
        if name in self.streams:
            return self.streams[name]
        s = Stream(name, self.reg,
                   mol=mol or {},
                   T=T if T is not None else self.default_T,
                   p=p if p is not None else self.default_p,
                   phase=phase)
        self.streams[name] = s
        return s

    def add_unit(self, unit: UnitOp) -> UnitOp:
        if unit.name not in self.units:
            self.units[unit.name] = unit
            self._unit_sequence.append(unit)
        return self.units[unit.name]

    def stream(self, name: str) -> Stream:
        if name not in self.streams:
            raise KeyError(f"Stream '{name}' not found in flowsheet")
        return self.streams[name]

    def run_once(self) -> None:
        """Execute all units in insertion order."""
        for u in self._unit_sequence:
            u.apply()

    def get_or_create_conditioned(
        self,
        src: Stream,
        *,
        target_T: Optional[float],
        target_p: Optional[float],
        do_condition: bool = True,
    ) -> Stream:
        """
        If do_condition is True AND the stream is not already at the target
        conditions, insert a Conditioner unit and return the conditioned stream.
        Otherwise return src unchanged.

        The conditioner unit and output stream are named deterministically:
          conditioner: "COND__{src.name}"
          output stream: "{src.name}__cond"
        Multiple calls with the same src are idempotent.
        """
        if not do_condition:
            return src

        need_T = (target_T is not None) and abs(src.T - target_T) > 1e-9
        need_p = (target_p is not None) and abs(src.p - target_p) > 1e-6

        # Check if this stream has already been conditioned to these targets
        cond_name = f"COND__{src.name}"
        out_name = f"{src.name}__cond"

        if cond_name in self.units:
            return self.streams[out_name]

        if not need_T and not need_p:
            return src

        # Create output stream
        out = self.add_stream(out_name, phase=src.phase, mol=dict(src.mol),
                              T=target_T if need_T else src.T,
                              p=target_p if need_p else src.p)

        # Create conditioner unit
        cond = Conditioner(cond_name,
                           target_T=float(target_T) if need_T else None,
                           target_p=float(target_p) if need_p else None)
        cond.connect_inlet("in", src)
        cond.connect_outlet("out", out)
        self.add_unit(cond)
        return out


# ─────────────────────────────────────────────────────────────────────────────
# RECYCLE SOLVER  (Anderson + under-relaxation)
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class RecycleResult:
    converged: bool
    iterations: int
    error: float
    error_history: List[float]


class _AndersonAccelerator:
    def __init__(self, m: int = 6, lam: float = 1e-10):
        self.m = int(m)
        self.lam = float(lam)
        self._X: List[np.ndarray] = []
        self._R: List[np.ndarray] = []

    def reset(self) -> None:
        self._X.clear(); self._R.clear()

    def propose(self, x_fp: np.ndarray, x_prev: np.ndarray) -> np.ndarray:
        r = x_fp - x_prev
        self._X.append(x_prev.copy())
        self._R.append(r.copy())
        if len(self._X) > self.m:
            self._X.pop(0); self._R.pop(0)
        k = len(self._R)
        if k < 2:
            return x_fp
        Rm = np.column_stack(self._R)
        RtR = Rm.T @ Rm + self.lam * np.eye(k)
        ones = np.ones((k, 1))
        KKT = np.block([[RtR, ones], [ones.T, np.zeros((1, 1))]])
        rhs = np.zeros((k + 1, 1)); rhs[-1] = 1.0
        try:
            c = np.linalg.solve(KKT, rhs)[:k, 0]
        except np.linalg.LinAlgError:
            return x_fp
        Fm = np.column_stack([self._X[j] + self._R[j] for j in range(k)])
        return Fm @ c


def _pack(fs: Flowsheet, names: List[str]) -> np.ndarray:
    return np.concatenate([fs.streams[n].to_dense() for n in names])


def _unpack(fs: Flowsheet, names: List[str], x: np.ndarray) -> None:
    nsp = fs.reg.n()
    for i, n in enumerate(names):
        chunk = x[i * nsp: (i + 1) * nsp]
        fs.streams[n].from_dense(np.maximum(chunk, 0.0))


def _rel_err(a: np.ndarray, b: np.ndarray, eps: float = 1e-12) -> float:
    denom = max(float(np.max(np.abs(b))), eps)
    return float(np.max(np.abs(a - b))) / denom


def solve_recycles(
    fs: Flowsheet,
    *,
    tears: Iterable[str],
    evaluate: Callable[[], None],
    max_iter: int = 300,
    tol: float = 1e-7,
    relax: float = 0.45,
    use_anderson: bool = True,
    anderson_m: int = 6,
    verbose: bool = True,
    print_every: int = 10,
) -> RecycleResult:
    """
    Solve recycle loops by successive substitution + Anderson acceleration.

    Parameters
    ----------
    fs        : Flowsheet
    tears     : names of tear streams (recycle streams whose state must converge)
    evaluate  : callable that runs the flowsheet once and updates tear streams
    """
    tear_names = [t for t in tears if t in fs.streams]
    if not tear_names:
        evaluate()
        return RecycleResult(True, 1, 0.0, [0.0])

    anderson = _AndersonAccelerator(m=anderson_m) if use_anderson else None
    x_prev = _pack(fs, tear_names)
    hist: List[float] = []

    for it in range(1, max_iter + 1):
        evaluate()
        x_fp = _pack(fs, tear_names)
        err = _rel_err(x_fp, x_prev)
        hist.append(err)

        if verbose and (it % print_every == 0 or it <= 3):
            print(f"  [recycle] iter {it:4d}  err = {err:.3e}")

        if err < tol:
            if verbose:
                print(f"  [recycle] CONVERGED at iter {it}, err = {err:.3e}")
            return RecycleResult(True, it, err, hist)

        # Anderson proposal
        x_prop = anderson.propose(x_fp, x_prev) if anderson else x_fp
        # Under-relaxation
        x_next = (1.0 - relax) * x_prev + relax * x_prop
        _unpack(fs, tear_names, x_next)
        x_prev = x_next

    if verbose:
        print(f"  [recycle] NOT CONVERGED after {max_iter} iters, err = {hist[-1]:.3e}")
    return RecycleResult(False, max_iter, hist[-1], hist)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

EXPORT_HIDDEN_STREAMS = {"F_X101", "F_X102", "F_X103", "F_StackGas", "U67_src", "F64_raw", "F64_aux", "F70_sink"}


def display_stream_name(name: str) -> str:
    """
    Convert internal compact stream labels (e.g. F24PT) to PFD-style dashed
    labels for user-facing output (e.g. F-24PT).
    """
    if name.startswith("F") and len(name) > 1 and name[1].isdigit():
        return "F-" + name[1:]
    return name


def export_to_excel(fs: Flowsheet, path: str) -> None:
    """
    Export stream table and unit summary to an Excel workbook.
    Sheets:
      Streams      – one row per stream: name, T, p, phase, F_total, m_dot
      MolarFlows   – one column per species
      MoleFractions
      Units        – one row per unit: name, class, inlet/outlet names
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not found — skipping Excel export")
        return

    wb = Workbook()

    # ── header style ─────────────────────────────────────────────────────────
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")

    def _hdr(ws, row: List):
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

    def _autofit(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

    visible_stream_names = [sname for sname in sorted(fs.streams) if sname not in EXPORT_HIDDEN_STREAMS]

    # ── Streams sheet ─────────────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Streams"
    _hdr(ws1, ["Stream", "T [K]", "p [Pa]", "Phase", "F_total [mol/s]", "m_dot [g/s]"])
    for sname in visible_stream_names:
        s = fs.streams[sname]
        ws1.append([display_stream_name(s.name), round(s.T, 4), round(s.p, 2), s.phase,
                    round(s.total_molar_flow(), 6), round(s.mass_flow_g_s(), 6)])
    _autofit(ws1)

    # ── Molar flows ───────────────────────────────────────────────────────────
    sp_list = list(fs.reg.species)
    ws2 = wb.create_sheet("MolarFlows")
    _hdr(ws2, ["Stream"] + sp_list)
    for sname in visible_stream_names:
        s = fs.streams[sname]
        ws2.append([display_stream_name(s.name)] + [round(s.get(sp), 8) for sp in sp_list])
    _autofit(ws2)

    # ── Mole fractions ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("MoleFractions")
    _hdr(ws3, ["Stream"] + sp_list)
    for sname in visible_stream_names:
        s = fs.streams[sname]
        Ft = s.total_molar_flow()
        fracs = [(s.get(sp) / Ft if Ft > EPS else 0.0) for sp in sp_list]
        ws3.append([display_stream_name(s.name)] + [round(f, 8) for f in fracs])
    _autofit(ws3)

    # ── Units sheet ───────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Units")
    _hdr(ws4, ["Unit", "Class", "Inlets", "Outlets", "Calcs"])
    for uname in [u.name for u in fs._unit_sequence]:
        u = fs.units[uname]
        ins = ", ".join(f"{k}={display_stream_name(v.name)}" for k, v in u.inlets.items())
        outs = ", ".join(f"{k}={display_stream_name(v.name)}" for k, v in u.outlets.items())
        calcs = "; ".join(f"{k}={_fmt(v)}" for k, v in (u.calcs or {}).items())
        ws4.append([u.name, u.__class__.__name__, ins, outs, calcs])
    _autofit(ws4)

    wb.save(path)
    print(f"  Exported to {path}")


def _fmt(v: Any) -> str:
    if isinstance(v, float):
        return f"{v:.4g}"
    return str(v)
